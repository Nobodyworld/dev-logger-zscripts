import argparse
import os
import warnings
from pathlib import Path

import openpyxl
from PIL import Image


def extract_image_details(image_path: Path):
    try:
        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("always")

            with Image.open(image_path) as img:
                width, height = img.size
                dpi = img.info.get("dpi", (0, 0)) if "dpi" in img.info else (0, 0)
                file_size_kb = image_path.stat().st_size / 1024  # size in KB

                # Convert the DPI values to integers
                horizontal_dpi = int(dpi[0])
                vertical_dpi = int(dpi[1])

                # If a DecompressionBombWarning was triggered, print the image name
                for warning in w:
                    if issubclass(warning.category, Image.DecompressionBombWarning):
                        print(
                            f"Warning: Image '{os.path.basename(image_path)}' "
                            f"might be a potential DOS bomb."
                        )

                # Return the details as a tuple
                return (
                    os.path.basename(image_path),
                    f"{width}x{height}",
                    f"{horizontal_dpi}x{vertical_dpi}",
                    file_size_kb,
                )

    except Exception as e:
        print(f"Error processing image '{os.path.basename(image_path)}': {e}")
        return (os.path.basename(image_path), "Error", "Error", "Error")


def create_excel(dir_path: Path, output_file: Path) -> None:
    # Create a new Excel workbook and get the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set header titles
    headers = ["Image Name", "Dimensions", "DPI", "File Size(KB)"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Populate the worksheet with image details
    row_num = 2
    for file in os.listdir(dir_path):
        if file.lower().endswith(".jpg"):
            details = extract_image_details(dir_path / file)
            for col_num, detail in enumerate(details, 1):
                ws.cell(row=row_num, column=col_num, value=detail)
            row_num += 1

    # Save the workbook
    wb.save(output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an Excel summary of JPEG dimensions and DPI."
    )
    parser.add_argument("input_dir", type=Path, help="Directory containing JPEG images.")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("image_details.xlsx"),
        # TODO - add global path function
        help="Output Excel file path (default: ./image_details.xlsx)",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    directory = args.input_dir.expanduser().resolve()
    if not directory.exists():
        raise FileNotFoundError(f"Input directory not found: {directory}")
    output = args.output.expanduser().resolve()
    create_excel(directory, output)


if __name__ == "__main__":
    main()
